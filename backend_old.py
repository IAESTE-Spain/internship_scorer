import tkinter
from tkinter import filedialog as fd
from tkinter import messagebox
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import json

# Data
allData = []
allIDs = []
allBidsConf = []
result = []
rules = None
CSVName = ''


def saveCSVName(fileCSV):
    global CSVName
    CSVName = fileCSV


##get all the ids
def getOffersIds(fileCSV):
    global allIDs
    with open(fileCSV, errors="ignore") as file_obj:
        reader_obj = csv.DictReader(file_obj, delimiter=';')
        for row in reader_obj:
            if "ES-20" in row['Ref.No']:
                allIDs.append(row['Ref.No'])
                allData.append(row)
    return allIDs


# Get the information of csv realted to one row
def getOfferInfo(id):
    return allData[allIDs.index(id)]


# Get the selected things for one bid already saved
def getOfferBidParameters(id):
    for bid in allBidsConf:
        if bid[0] == id:
            return bid


# Find if there is already a row with that id that is saved and if not just adds the row to config
def setOfferBidParameters(row):
    for index, line in enumerate(allBidsConf):
        if line[0] == row[0]:
            allBidsConf[index] = row
            return
    allBidsConf.append(row)


# Start of methods for score different parts needed
def submitedOfferConfig():
    return rules.get('base')


def checkTimeDeliveryConfig(isNotOnTime):
    return rules.get('term') if isNotOnTime else 0


# Si pone OR se utiliza la que menos perjudique Y si es and se suman directamente. AND + (I OR II)
def languagesConfig(row):
    lanRules = rules.get('language')
    result = 0
    allResults = []
    isSpanish = False

    for option in lanRules['check']:
        lang = row[option]
        if lang:

            if option[-1].isdigit() and lang != 'Spanish':
                if lang in lanRules:
                    allResults.append(lanRules[lang])
                else:
                    allResults.append(lanRules['default'])
            elif option[-1].isdigit() and lang == 'Spanish':
                return
            else:
                return

    return result


""" def languagesConfig(row):
    lanRules = rules.get('language')
    result = 0
    allresutlts=[]
    isSpanish = 0
    checkOrAdd = 0

    for opt in lanRules['check']:
        lang = row[opt]
        if lang != "":
            if opt[-1].isdigit():
                if lang != 'Spanish':
                    if lang in lanRules:
                        allresutlts.append(lanRules[lang])
                    else:
                        allresutlts.append(lanRules['default'])

                else:
                    isSpanish = 1
            else:
                if 'Level' in opt and isSpanish == 1:
                    langFormated = lang.split('(')[0].split()
                    print("formated",langFormated[0])
                    print("original",lang)
                    allresutlts.append(lanRules['Spanish'][langFormated[0]])
                    isSpanish = 0
                elif 'Level' not in opt:
                    allresutlts.append(row[opt])

    for option in allresutlts:
        if isinstance(option, int):
            if checkOrAdd == 1:
                    result += option
            else:
                if result < option:
                    result = option
                else:
                    if option == 'Or':
                        checkOrAdd = 0
                    else: 
                        checkOrAdd = 1


    return result """


def languagesConfig(row):
    lanRules = rules.get('language')
    result = 0
    allresults = []
    isSpanish = False

    # Primera fase: recopilación de resultados basados en las reglas de idiomas.
    for opt in lanRules['check']:
        lang = row.get(opt, "")
        if lang:
            if opt[-1].isdigit() and lang != 'Spanish':
                allresults.append(lanRules.get(lang, lanRules['default']))
            elif lang == 'Spanish':
                isSpanish = True
            else:
                if 'Level' in opt and isSpanish:
                    langFormatted = lang.split('(')[0].split(maxsplit=1)[0]
                    print("/./////////////////////////////////////////")
                    print(lang)
                    print(langFormatted)
                    allresults.append(lanRules['Spanish'].get(langFormatted, 0))
                    isSpanish = False
                elif 'Level' not in opt:
                    allresults.append(lanRules.get(lang, 0))

    # Segunda fase: cálculo de la puntuación con lógica AND y OR.
    and_sum = 0
    or_values = []
    for value in allresults:
        if isinstance(value, int):
            and_sum += value  # Sumamos todos los valores directamente, asumiendo que son ANDs.
        elif value == 'Or':
            if or_values:  # Si ya hay valores OR, elegimos el menor y lo sumamos a and_sum.
                and_sum += min(or_values)
                or_values.clear()  # Limpiamos los valores OR para empezar de nuevo.
        else:  # Aquí manejamos los valores que no son enteros ni 'Or', si los hubiera.
            pass  # Puedes agregar lógica adicional aquí según sea necesario.

    # Si quedan valores OR después del último ciclo, los procesamos.
    if or_values:
        and_sum += min(or_values)

    return and_sum


# Si el inicio y el final incluye julio y agosto tambien cuenta como 40 puntos.
def periodConfig(row):
    result = 0
    minWeeks = int(row[rules["period"]["checkWeeks"][0]])
    maxWeeks = int(row[rules["period"]["checkWeeks"][1]])
    startMonth = int(row[rules["period"]["checkMonth"][0]].split("-")[1])
    endMonth = int(row[rules["period"]["checkMonth"][1]].split("-")[1])

    if minWeeks <= 8 or maxWeeks <= 8:
        if startMonth == 7 and endMonth == 8:
            result = rules["period"]["8"]
        return result

    if minWeeks <= 12 or maxWeeks <= 12:
        if startMonth <= 7 and endMonth >= 8:
            result = rules["period"]["8"]
            return result
        result = rules["period"]["12"]
        return result

    result = rules["period"]["default"]
    return result


def salaryConfig(row):
    result = 0

    frequency = row[rules['payment']['check'][1]]
    pay = float(row[rules['payment']['check'][0]])
    stepAddingPoints = rules['payment']['over']
    minPaymentMonth = rules['payment']['minMonth']

    if frequency != 'Monthly':
        pay = pay * rules['payment']['week']

    # Tener en cuenta el alojamiento y comida, si tienen hay que sumar el valor al salario y contar los punto en caso de superar 800
    if pay <= minPaymentMonth:
        return 0

    # Que sea int redondendo a la alta siempre que sea mayor a 1 si es menor a 1 se redondea a la baja
    result = (pay - minPaymentMonth) // stepAddingPoints

    return result * rules['payment']['points']


# Pais se supone que tendria que estar bien ya que se selecciona por el usuario
# Cambiar texto Union europea add Otan
def countryrestrictionsConfig(restriction):
    return rules['country'][restriction] if restriction != 'Sin restricción' else 0


def findPos(discipline):
    allDisciplines = rules['generalDisciplines']['disciplines']
    i = 0
    j = 0

    while i < len(allDisciplines):
        if discipline != allDisciplines[i]['field'].lower():
            j = 0
            while j < len(allDisciplines[i]['FieldsOfStudy']):
                if discipline == allDisciplines[i]['FieldsOfStudy'][j]['field']:
                    return allDisciplines[i]['FieldsOfStudy'][j]
                j += 1
        else:
            return allDisciplines[i]

        i += 1

    return 0


def fosConfig(row):
    result = 0
    generalDiscipline = row[rules['generalDisciplines']['check'][0]]
    fieldsofstudy = row[rules['generalDisciplines']['check'][1]].split(";")

    if "|" in generalDiscipline:
        generalDiscipline = generalDiscipline.split("|")

    if isinstance(generalDiscipline, str):
        position = findPos(generalDiscipline.lower())
        if position != 0:
            if result <= position['points']:
                result = position['points']
    else:
        for ds in generalDiscipline:
            position = findPos(ds.lower())
            if position != 0:
                if result <= position['points']:
                    result = position['points']

    if len(fieldsofstudy):
        for fs in fieldsofstudy:
            position = findPos(fs.lower())
            if position != 0:
                if result <= position['points']:
                    result = position['points']

    return result


# Make all the correct points with for the different rows
def startCheckingRules():
    global result
    global rules

    rules = 'assets/Reglas.json'
    with open(rules, 'r') as archivo:
        rules = json.load(archivo)

    print("start checking")
    result.append(
        ['Ref.No', 'Total', 'Practica entregada', 'Plazo de entrega', 'Idiomas', 'Duración y tipo de periodo', 'Sueldo',
         'Países', 'Valoración especialidad'])

    for offer in allBidsConf:
        resultRow = []
        row = getOfferInfo(offer[0])
        resultRow.append(offer[0])
        resultRow.append(0)

        submitedOffer = submitedOfferConfig()
        resultRow.append(submitedOffer)

        deliveryTime = checkTimeDeliveryConfig(offer[2])
        resultRow.append(deliveryTime)

        languages = languagesConfig(row)
        resultRow.append(languages)

        period = periodConfig(row)
        resultRow.append(period)

        salary = salaryConfig(row)
        resultRow.append(salary)

        country = countryrestrictionsConfig(offer[4])
        resultRow.append(country)

        fos = fosConfig(row)
        resultRow.append(fos)

        result.append(resultRow)

    print(result)
    getResultBid()
    pass


# Saves all the information in the csv file, it needs the format also will break in the last for and it need and end windows(there is a finish method)
def saveConfigToPoints():
    resultdocument = "ResultOffers"
    with open(resultdocument, 'w', ) as file2_obj:
        writer_obj = csv.writer(file2_obj)
        for row in result:
            writer_obj.writerow(row)


# Get the final result of the offer, and save it in the array result
def getResultBid():
    global result

    for i in range(1, len(result)):
        row = result[i]
        resultSum = row[2:]
        resultNum = 0
        for element in resultSum:
            resultNum += element

        row[1] = resultNum
    createCSV()


def createCSV():
    global CSVName
    # Crear un libro de trabajo y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active

    # Añadir los datos al libro de trabajo
    for row_data in result:
        ws.append(row_data)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Formatear la cabecera (primera fila)
    for row in ws.iter_rows(min_row=1, max_row=len(result), min_col=1, max_col=len(result[0])):
        for cell in row:
            # a
            # Aplicar bordes
            cell.border = thin_border
            # Formato para la cabecera
            if cell.row == 1:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Color amarillo
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            # Formato para la segunda columna
            if cell.column == 2:
                cell.font = Font(color="FF0000", bold=True)

    # Ajustar el tamaño de las celdas para que se acomoden al contenido
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:  # Necesario para manejar valores no string
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Obtener la ruta del escritorio del usuario
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    print(CSVName)
    # Nombre del archivo de Excel
    nombre_archivo = "RESULT_" + CSVName + ".xlsx"

    # Ruta completa para guardar el archivo
    ruta_completa = os.path.join(desktop, nombre_archivo)

    # Guardar el libro de trabajo
    wb.save(ruta_completa)

    print(f"Archivo guardado en: {ruta_completa}")
