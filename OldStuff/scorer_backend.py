"""Module used for checking scores"""
import csv
import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import time

global csv_manager
global bid_manager
global rules_manager


def get_all_offers_id() -> list[str]:
    """Returns all the ids from all the offers saved"""
    return csv_manager.data_ids


def get_offer_info(offer_id: str) -> list[str]:
    """Returns all the information related to the offer given that ID"""
    return csv_manager.data[csv_manager.data_ids.index(offer_id)]


def get_bid_parameters(offer_id: str) -> list[str]:
    """Return configuration bid saved"""
    if offer_id not in bid_manager.bids:
        raise KeyError("Bid not saved")
    return bid_manager.bids.get(offer_id)


def save_bid_parameters(offer_id: str, parameters: list[str]) -> None:
    """Save the parameters of a bid"""
    bid_manager.bids[offer_id] = parameters


def is_delivered_on_time(delivered_on_time: bool) -> bool:
    """The offer was delivered on time?"""
    return rules_manager.on_term if delivered_on_time else rules_manager.past_term


def points_for_languages(row: list[str]) -> int:
    """Points related to language"""
    selected_results: list[tuple[int, str]] = []
    actual_language_points: int = 0
    actual_language: str = ""
    is_spanish: bool = False
    result: int = 0
    for option in rules_manager.languages_options_check_in_csv:
        checking_language: str = row[option]
        if not checking_language:
            continue

        if "Level" not in option and "Or" not in option:
            if checking_language == 'Spanish':
                actual_language = checking_language
                is_spanish = True
                continue

            actual_language = checking_language
        elif "Level" in option and is_spanish:
            spanish_check: str = actual_language.join(checking_language)
            actual_language_points = rules_manager.languages_points.get(spanish_check)
        elif "Level" in option and not is_spanish:
            if actual_language not in rules_manager.languages_points:
                actual_language_points = rules_manager.languages_points.get("Default")
            else:
                actual_language_points = rules_manager.languages_points.get(actual_language)
        elif "Or" in option:
            selected_results.append((actual_language_points, checking_language))
            actual_language_points = 0
            is_spanish = False
            actual_language = ""

    for language_points, operation_language in selected_results:
        result = result + language_points if operation_language == "AND" else max(result, language_points)

    return result


def points_for_duration_offer(row: list[str]) -> int:
    """Points for the duration of the offer"""
    minimum_weeks: int = int(row[rules_manager.period_weeks_check[0]])
    maximum_weeks: int = int(row[rules_manager.period_weeks_check[1]])
    starting_month: int = int(row[rules_manager.period_month_check[0]])
    ending_month: int = int(row[rules_manager.period_month_check[1]])

    if (minimum_weeks <= 8 or maximum_weeks <= 8) and starting_month == 7 and ending_month == 8:
        return rules_manager.points_weeks.get("8")

    if minimum_weeks <= 12 or maximum_weeks <= 12:
        if starting_month <= 7 and ending_month >= 8:
            return rules_manager.points_weeks.get("8")
        return rules_manager.points_weeks.get("12")

    return rules_manager.points_default.get("8")


def points_for_salary(row: list[str]) -> int:
    """Points to give for the salary of the offer"""
    payment_amount = float(row[rules_manager.salary_options_check[0]])
    payment_frequency: str = row[rules_manager.salary_options_check[1]]

    if payment_frequency != 'Monthly':
        payment_amount = payment_amount * rules_manager.salary_minimum_pay_per_month

    if payment_frequency <= rules_manager.salary_minimum_pay_per_month:
        return 0

    result_salary: float = (
                                   payment_amount - rules_manager.salary_minimum_pay_per_month) // rules_manager.salary_step_amount

    return result_salary * rules_manager.salary_point_per_step


def points_per_country_restriction(restriction_country: str) -> int:
    """Points related to the country restriction"""
    return rules_manager.country_points.get(restriction_country) if restriction_country != 'Sin restricción' else 0


def points_per_general_discipline(discipline: str) -> int:
    """Points given by the general discipline"""
    return rules_manager.disciplines_options.get(discipline, {}).get("points", 0)


def points_per_field_of_study(discipline: str, fos: str) -> int:
    """Points given by the field of study"""
    return rules_manager.disciplines_options.get(discipline, {}).get("FieldsOfStudy", {}).get(fos, 0)


def get_points_of_each_parameter_of_offer(row: list[str]) -> None:
    general_discipline: int = points_per_general_discipline("")
    field_of_study: int = points_per_field_of_study("", "")

    points_offer: list[any] = [rules_manager.base_points, is_delivered_on_time(True), points_for_languages(row),
                               points_for_duration_offer(row),
                               points_for_salary(row), points_per_country_restriction(row[4]),
                               max(general_discipline, field_of_study)]

    result_points: int = get_total_points_for_each_offer(points_offer)
    points_offer.insert(0, result_points)
    points_offer.insert(0, row[0])
    csv_manager.result.append(points_offer)
    ...


def get_total_points_for_each_offer(offer_points: list[int]) -> int:
    """Check the point for an offer"""
    return sum(offer_points)


def create_csv_file() -> None:
    """Generates the CSV file and stores it in the computer."""

    # Creat work book and select active sheet
    workbook = Workbook()
    wordbook_sheet = workbook.active

    # Add data to the active sheet
    for row in bid_manager.result:
        wordbook_sheet.append(row)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Formating first row
    for row in wordbook_sheet.iter_rows(min_row=1, max_row=len(bid_manager.result), min_col=1,
                                        max_col=len(bid_manager.result[0])):
        for cell in row:
            # Apply borders
            cell.border = thin_border
            # Format headers
            if cell.row == 1:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Color amarillo
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            # Format second column
            if cell.column == 2:
                cell.font = Font(color="FF0000", bold=True)

    # Adjust the size of cells to fit the content
    for col in wordbook_sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Obtain the letter of the column
        for cell in col:
            try:  # Necessary to manage values and not strings
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                print("error with columns")
                pass

        wordbook_sheet.column_dimensions[column].width = (max_length + 2)

    # Obtain path
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    print(csv_manager.csv_name)
    # Name of the Excel file
    file_name = "RESULT_" + csv_manager.csv_name + "_" + time.strftime("%Y%m%d-%H%M%S") + ".xlsx"

    # Full path
    full_path = os.path.join(desktop, file_name)

    # Save the active sheet
    wordbook_sheet.save(full_path)

    print(f"File saved: {full_path}")


class CSVManager:
    """Class that saves all the data related to the CSV"""

    def __init__(self, csv_file: str):
        """Initializes the CSV manager"""
        self.csv_file: str = csv_file
        self.csv_name: str = csv_file.split("/")[-1].split(".")[0]
        self.data: list[dict[str, str]] = self.load_data()
        self.data_ids: list[str] = [row['Ref.No'] for row in self.data if "ES-20" in row['Ref.No']]

    def load_data(self) -> list[dict[str, str]]:
        """Loads all the data for an easy access"""
        with open(self.csv_file) as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            # check here the option of the delimiter to avoid problems
            return [row for row in reader]


class BidManager:
    """Manage the information related to the offers saved"""

    def __init__(self, total_offers: int):
        """Initializes the offers manager"""
        self.total_offers: int = total_offers
        self.bids: dict[str, list[str]] = {}
        self.result: list[any] = ['Ref.No', 'Total', 'Practica entregada', 'Plazo de entrega', 'Idiomas',
                                  'Duración y tipo de periodo', 'Sueldo', 'Países', 'Valoración especialidad']


class RulesManager:
    """Manager for the rules used to process the offers"""

    def __init__(self):
        """Initializes the rules manager"""
        with open('assets/Reglas.json', 'r') as rules:
            rules = json.load(rules)
        self.base_points: int = rules['base']
        self.on_term: int = rules['on_term']
        self.past_term: int = rules['past_term']
        language_code: str = 'language'
        self.languages_points: dict[str, int] = rules[language_code]['languages']
        self.languages_options_check_in_csv: list[str] = rules[language_code]['check']
        period_code: str = 'period'
        self.period_weeks_check: list[str] = rules[period_code]['checkWeeks']
        self.period_month_check: list[str] = rules[period_code]['checkMonth']
        self.points_weeks: list[str] = rules[period_code]['pointsWeeks']
        self.points_default: list[str] = rules[period_code]['default']
        salary_code: str = 'payment'
        self.salary_options_check: list[str] = rules[salary_code]['check']
        self.salary_step_amount: int = rules[salary_code]['stepAmount']
        self.salary_minimum_pay_per_week: int = rules[salary_code]['minWeek']
        self.salary_minimum_pay_per_month: int = rules[salary_code]['minMonth']
        self.salary_week_equal_month: float = rules[salary_code]['week']
        self.salary_point_per_step: int = rules[salary_code]['points']
        country_code: str = 'country'
        self.country_check: list[str] = rules[country_code]['check']
        self.country_points: dict[str, int] = rules[country_code]['countries']
        disciplines_code: str = 'discipline'
        self.disciplines_check: list[str] = rules[disciplines_code]['check']
        self.disciplines_options: dict[str, dict[str, any]] = rules[disciplines_code]['disciplines']


def main() -> None:
    """Main method"""
    ...


if __name__ == '__main__':
    main()


def initializer_file(csv_file: str) -> None:
    """Initializer managers"""
    global csv_manager
    global bid_manager
    global rules_manager

    csv_manager = CSVManager(csv_file)
    bid_manager = BidManager(len(csv_manager.data))
    rules_manager = RulesManager()
